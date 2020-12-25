import time
import random
import pandas as pd
from selenium import webdriver
from openpyxl import load_workbook
from datetime import datetime, timedelta
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

df = pd.read_excel('excel.xlsx')
data = None
chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9014")
#Change chrome driver path accordingly
chrome_driver = "C:\Program Files (x86)\chromedriver.exe"
driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)

def updateInsuranceNo(n):
    try:
        driver.implicitly_wait(100)
        #insurance = driver.find_element_by_id("ctl00_HomePageContent_ctrlLabelIPNumber").text
        filename = "excel.xlsx"
        wb = load_workbook(filename)
        ws = wb.worksheets[0]
        row = f'J{n}'
        ws[row] = n*100
        wb.save(filename)
        time.sleep(3)
    except Exception as e:
        print(e)
    #driver.find_element_by_id("ctl00_HomePageContent_close").click()
    



def getDate():
    N_DAYS_AGO = 9
    today = datetime.now()    
    n_days_ago = today - timedelta(days=N_DAYS_AGO)
    return n_days_ago.strftime('%d-%m-%Y').replace('-', '/')
    
def isMarried():
    dob = int(data['DOB'][6:10])
    if dob <= 1994:
        return 1
    else:
        return 0

def getAddress():
    address = ['Malakpur', 'Chhalera', 'Surajpur', 'Sadarpur', 'Bhangel','Dadri','Chouganpur',"Nagla","Phase 2"]
    return random.choice(address)

def mobile_registration(name):
    def getMobile_no(number):
        return str(number - random.randint(10,90))
    try:
        
        driver.find_element_by_id("ctl00_HomePageContent_rbtnlistIsregistered_1").click()
        driver.implicitly_wait(10)
        obj1 = driver.switch_to.alert
        obj1.accept()
        
        
        driver.implicitly_wait(10)
        mobile = driver.find_element_by_id("ctl00_HomePageContent_txtmobilenumber")
        mobile.send_keys(str(data['MOBILE NO']))
        driver.find_element_by_link_text("Validate Mobile Number").click()
        time.sleep(2)
        driver.find_element_by_id("ctl00_HomePageContent_btnContinue").click()
        time.sleep(2)
        obj1 = driver.switch_to.alert
        obj1.accept()
    except:
        print(f"{name} is registered already")
        driver.implicitly_wait(10)
        driver.find_element_by_id("ctl00_HomePageContent_btncancel").click()
        time.sleep(3)
        mobile = driver.find_element_by_name("ctl00$HomePageContent$txtmobilenumber")
        mobile.send_keys(getMobile_no(data['MOBILE NO']))
        driver.find_element_by_link_text("Validate Mobile Number").click()
        time.sleep(3)
        #due to element click intercepted
        javascript = "document.getElementById('ctl00_HomePageContent_btnContinue').click()"
        driver.execute_script(javascript)
        #driver.find_element_by_id("ctl00_HomePageContent_btnContinue").click()
        obj = driver.switch_to.alert
        obj.accept()
        driver.implicitly_wait(60)
    
         



def employee_registration():
    Name_employee = driver.find_element_by_id("ctl00_HomePageContent_ctrlTextEmpName")
    Name_employee.clear()
    Name_employee.send_keys(data['NAME'])

    Father_name = driver.find_element_by_id("ctl00_HomePageContent_ctrlTextFatherHusName")
    Father_name.clear()
    Father_name.send_keys(data['FATHER'])

    javascript = f"document.getElementById('ctl00_HomePageContent_ctrlTxtIpDate').value='{str(data['DOB'])}'"
    driver.execute_script(javascript)

    javascript = f"document.getElementById('ctl00_HomePageContent_ctrlDIDateOfAppointmentDy').value='{getDate()}'"
    driver.execute_script(javascript)

    select_Maritial = Select(driver.find_element_by_id('ctl00_HomePageContent_ctrlRDMarried'))
    select_Maritial.select_by_index(isMarried())

    employee_address = driver.find_element_by_id("ctl00_HomePageContent_ctrlTextPresentAddress1")
    employee_address.clear()
    employee_address.send_keys(getAddress())

    selectState = Select(driver.find_element_by_id('ctl00_HomePageContent_ctrlTxtPresentState'))
    selectState.select_by_visible_text("Uttar Pradesh")

    selectDistrict = Select(driver.find_element_by_id('ctl00_HomePageContent_ctrlTextPresentDistrict'))
    selectDistrict.select_by_visible_text("Gautam Buddha Nagar")

    driver.find_element_by_id("ctl00_HomePageContent_chkboxCopyPresentAddress").click() #copy the address

    selectDispensaryState = Select(driver.find_element_by_id('ctl00_HomePageContent_ddlDispensaryState'))
    selectDispensaryState.select_by_visible_text("Uttar Pradesh")
    try:
        selectDispensaryDistrict = WebDriverWait(driver, 60).until(
        EC.text_to_be_present_in_element((By.ID, 'ctl00_HomePageContent_ddlDispensaryDistrict'), "Gautam Buddha Nagar")
        )

    finally:
        selectDispensaryDistrict = Select(driver.find_element_by_id('ctl00_HomePageContent_ddlDispensaryDistrict'))
        selectDispensaryDistrict.select_by_visible_text("Gautam Buddha Nagar")

    try:
        selectTextDispensary = WebDriverWait(driver, 60).until(
        EC.text_to_be_present_in_element((By.ID, 'ctl00_HomePageContent_ctrlTextDispensary'), "Greater Noida, UP (ESIC Disp.)")
        )

    finally:
        selectTextDispensary = Select(driver.find_element_by_id('ctl00_HomePageContent_ctrlTextDispensary'))
        selectTextDispensary.select_by_visible_text("Greater Noida, UP (ESIC Disp.)")

    selectDispensaryState = Select(driver.find_element_by_id('ctl00_HomePageContent_ddldependantDispensaryState'))
    selectDispensaryState.select_by_visible_text("Uttar Pradesh")
    try:
        selectDispensaryDistrict = WebDriverWait(driver, 60).until(
        EC.text_to_be_present_in_element((By.ID, 'ctl00_HomePageContent_ddldependantDispensaryDistrict'), "Gautam Buddha Nagar")
        )

    finally:
        selectDispensaryDistrict = Select(driver.find_element_by_id('ctl00_HomePageContent_ddldependantDispensaryDistrict'))
        selectDispensaryDistrict.select_by_visible_text("Gautam Buddha Nagar")

    try:
        selectTextDispensary = WebDriverWait(driver, 60).until(
        EC.text_to_be_present_in_element((By.ID, 'ctl00_HomePageContent_ddldependantdispensary'), "Greater Noida, UP (ESIC Disp.)")
        )

    finally:
        selectTextDispensary = Select(driver.find_element_by_id('ctl00_HomePageContent_ddldependantdispensary'))
        selectTextDispensary.select_by_visible_text("Greater Noida, UP (ESIC Disp.)")

    nomineeDetailsLink = driver.find_element(By.XPATH, '//tr[@id="Tr11"]/td[last()]/a')
    nomineeDetailsLink.click()

    bankDetailsLink = driver.find_element(By.XPATH, '//tr[@id="Tr18"]/td[last()]/a')
    bankDetailsLink.click()




def nominee_details(nomineeDetailsLink):
    driver.switch_to.window(nomineeDetailsLink)
    driver.implicitly_wait(10)
    father_nominee = driver.find_element_by_id("ctl00_HomePageContent_ctrlTextUserName")
    father_nominee.clear()
    father_nominee.send_keys(data["FATHER"])

    relationship = Select(driver.find_element_by_id('ctl00_HomePageContent_RelationShipWithIp'))
    relationship.select_by_index(7)

    adress_nominee = driver.find_element_by_id("ctl00_HomePageContent_ctrlTextAddress1")
    adress_nominee.clear()
    adress_nominee.send_keys("DO")

    select_nominee_State = Select(driver.find_element_by_id('ctl00_HomePageContent_States'))
    select_nominee_State.select_by_visible_text("Uttar Pradesh")
    try:
        select_nominee_District = WebDriverWait(driver, 60).until(
        EC.text_to_be_present_in_element((By.ID, 'ctl00_HomePageContent_Districts'), "Gautam Buddha Nagar")
        )

    finally:
        select_nominee_District = Select(driver.find_element_by_id('ctl00_HomePageContent_Districts'))
        select_nominee_District.select_by_visible_text("Gautam Buddha Nagar")
    driver.find_element_by_id("ctl00_HomePageContent_rbtnlistNomneeAkaFamily_0").click()
    driver.find_element_by_id("ctl00_HomePageContent_Save").click()
    try:
        element = WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.ID, "ctl00_HomePageContent_btnClose"))
        )
    finally:
        time.sleep(2)
        driver.find_element_by_id("ctl00_HomePageContent_btnClose").click()



def bank_details(bankDetailsLink,employeeRegistration):
    def generateAC(account_no):
        return str(account_no- random.randint(1,1000))

    driver.switch_to.window(bankDetailsLink)
    Ifsc_detail = driver.find_element_by_id("ctl00_HomePageContent_txtIFSCcode")
    Ifsc_detail.clear()
    Ifsc_detail.send_keys(data["IFSC"])
    driver.find_element_by_id("ctl00_HomePageContent_btnIFSCcode").click()
    # driver.implicitly_wait(20)
    try:
        element = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "ctl00_HomePageContent_txtacc_number"))
        )
    finally:
        try:
            account_type = Select(driver.find_element_by_id('ctl00_HomePageContent_ddlAccountType'))
            account_type.select_by_index(3)
            account_no = driver.find_element_by_id("ctl00_HomePageContent_txtacc_number")
            account_no.send_keys(str(data['AC']))
            driver.find_element_by_id("ctl00_HomePageContent_btnsubmit").click()
            
            try:
                WebDriverWait(driver, 3).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                alert.accept()
                driver.implicitly_wait(10)
                account_no = driver.find_element_by_id("ctl00_HomePageContent_txtacc_number")
                account_no.send_keys(generateAC(data['AC']))
                driver.find_element_by_id("ctl00_HomePageContent_btnsubmit").click()
                time.sleep(3)
                driver.find_element_by_id("btnCancel").click()
            except:
                time.sleep(3)
                driver.find_element_by_id("btnCancel").click()
        finally:
            driver.switch_to.window(employeeRegistration)

        
def main_work(i):
    try:
        global data
        data = df.loc[i]
        print(f"Processing ---{data['NAME']} ---")
        parent_handle = driver.current_window_handle
        driver.find_element_by_xpath("//*[@id='lnkRegisterNewIP']").click()
        handle = driver.window_handles
        driver.switch_to.window(handle[1])
        mobile_registration(data['NAME'])
        employee_registration()
        handles = driver.window_handles
        size = len(handles)
        employeeRegistration = handles[1]
        nomineeDetailsLink = handles[3]
        bankDetailsLink = handles[2]
        nominee_details(nomineeDetailsLink)
        bank_details(bankDetailsLink,employeeRegistration)
        driver.implicitly_wait(20)
        driver.find_element_by_id("ctl00_HomePageContent_dec_chkbox").click()
        # driver.find_element_by_id("ctl00_HomePageContent_Submit").click()
        updateInsuranceNo(i+2)
        handles = driver.window_handles
        size = len(handles)
        for x in range(size):
            if handles[x] != parent_handle:
                driver.switch_to.window(handles[x])
                driver.close()
        driver.switch_to.window(parent_handle)
        time.sleep(5)
        print(f"Succesful --- {data['NAME']} ---")
    except Exception as e:
        print(f"Failure --- {data['NAME']} --- due to {e}")
        handles = driver.window_handles
        size = len(handles)
        for x in range(size):
            if handles[x] != parent_handle:
                driver.switch_to.window(handles[x])
                driver.close()
        driver.switch_to.window(parent_handle)
        return

def automatically():
    for index in range(df.index.stop):
        try:
            main_work(index)
        except Exception as e:
            print(e)
            continue

def manually():
    n = int(input("Enter value of employee: "))
    if (n < 2):
        print("Not possible")
        return
    else:
        try:
            main_work(n-2)
        except Exception as e:
            print(e)
            return

if __name__ == '__main__':
    print("How do you want to operate :\n1.Automatically  or 2.Manually")
    choice = int(input("Enter valid choice: "))
    if choice == 1:
        automatically()
    elif choice == 2:
        manually()
    else:
        print("Wrong choice")
    
        
            
    
    
    




   